import streamlit as st
import openpyxl
from openpyxl.styles import Font
import random
from datetime import datetime, timedelta
import io

# Template data from document (split and padded to exactly 84 columns)
row1_text = "DELETE row before importing,PERSON DETAILS TAB - Yellow fields are mandatory,,,,,,,,,,,,,,,,,,,,PERSON CONTACT DETAILS - INCLUDES EMERGENCY CONTACTS,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,PERSON TRAINING DETAILS,,PERSON ROSTER DETAILS,,,,,PERSON TRAVEL SETTINGS,,PERSON CAMP DETAILS,,,,,COST CODE,PERSON TRAVEL OPTIONS,,,,,,,,,,,,,CUSTOM FIELDS on Person Details Tab (can import up to 20 custom fields),,,"
row2_text = "DELETE row before importing,Username is an Email Address - generally the persons Work Email. Username is the unique identifier used for logging in to PeopleTray,,SYSTEM OPTIONS IN PT: Mr; Mrs; Ms; Miss; Dr,Use CAPITALS for Last name,All People Must be in a Workgroup - Used to define WG Editors etc,,,,Workroles are used in PT to manage Compliance. People can have multiple Workroels,Additional Workrole,Additional Workrole,Job Titles are different to Workroles,Area is used in CRMTray, Projects & TeamTray - Example South of River; Interstate; International,Example - FullTime; Part Time; Casual; Contractor,Using Employer Selector,Using CRMTray Companies,SYSTEM OPTIONS: Male; Female; Unspecified,35442,44927,45292,,Must use International Format - '+61444999888,All notifications will be sent to this email,Some notifications will be sent to this email eg. Travel Notifications,,,,,,,,,,,,,,,,Person First Emergency Contact details,,,,,,Person Second Emergency Contact details,,,,,,Displays on Training Tab,Displays on Training Tab,Name of Roster Cycle Configured in PeopleTray,,,,,Where Person regularly Travels from,Where Person regularly Travels to,Camp Person regularly stays,Used to define Block at Camp Person regularly stays,Used to define Room Type Camp Person regularly stays,Permanent Room Assigned,Permanent Room Group  Assigned,Single Cost code assigned on Roster Tab,Number 1 to 25,Number 1 to 25,,Used where person is above or below default weight,Used where person regularly carries extra baggage for Travel bookings,(Label can be renamed),(Label can be renamed),Yes or No,Yes or No,Yes or No,Yes or No,Include details of any Other Requirements,Yes or No,Must be configured in PT before import,Must be configured in PT before import,Must be configured in PT before import,Must be configured in PT before import"
row3_text = "DO NOT DELETE,Username,Employee ID,Salutation,Last Name,First Name,Middle Name,Preferred Name,Workgroup,Work Role 01,Work Role 02,Work Role 03,Job Title,Area,Employment Type,Employer,Person Company,Gender,Date of Birth,Date Commenced,Date Ceased,Work Phone,Work Mobile,Email,Additional Email,Home Phone,Home Mobile,Home Email,Address 1,Address 2,Suburb,Postcode,State,Country,Postal Address 1,Postal Address 2,Postal Suburb,Postal Postcode,Postal State,Postal Country,EmergencyContact - 1 Name,EmergencyContact - 1 Phone,EmergencyContact - 1 Email,EmergencyContact - 1 Relationship,EmergencyContact - 1 Guardian,EmergencyContact - 1 Can Speak on Behalf of,EmergencyContact - 2 Name,EmergencyContact - 2 Phone,EmergencyContact - 2 Email,EmergencyContact - 2 Relationship,EmergencyContact - 2 Guardian,EmergencyContact - 2 Can Speak on Behalf of,USI,Default RTO,Current Roster Name,Current Roster From Date,Person Roster Seed Date,Current Roster To Date,Current Roster Notes,Home Port,Default Site Port,Default Camp,Default Block,Default Room Type,Default Room,Default Room Group,Default Cost Code,Inbound Preference,Outbound Preference,Seat Preference,Weight KG,Extra Baggage KG,Frequent Flyer 1,Frequent Flyer 2,Vegetarian,Vegan,Gluten Free,Other Requirement,Other Requirement Note,Wheelchair Access Required,Custom Field 01,Custom Field 02,Custom Field 03,Custom Field 04"

# Split and pad to 84 columns
def prepare_row(text):
    row = [cell.strip() for cell in text.split(',')]
    while len(row) < 84:
        row.append('')
    return row[:84]  # Trim if extra

row1 = prepare_row(row1_text)
row2 = prepare_row(row2_text)
row3 = prepare_row(row3_text)

# Default first names (hardcoded 200-name list)
first_names_list = ["Charlotte","Evelyn","Willow","Ariana","Isabella","Michelle","Betty","Gabriella","Victoria","Anna","Laura","Eva","Riley","Kennedy","Carol","Scott","Sofia","Susan","Lucy","Ava","Jack","Michelle","Willow","Zoey","Frank","Sarah","David","George","Robert","Patrick","Sarah","Alice","Gary","Matthew","Isla","Brenda","Lisa","Mark","Caroline","Joseph","Ava","Joshua","Skylar","Zoey","Eliana","Alice","Karen","Barbara","Lily","Justin","Mary","Joshua","Serenity","Benjamin","Thomas","Brian","Angela","Alice","Serenity","Eva","Jonathan","Madelyn","Richard","Jennifer","Laura","Adeline","Emilia","Gianna","Willow","Amelia","Sophia","Christopher","Kevin","Alice","Mark","Emma","Claire","Ella","Savannah","Stella","Matthew","Emma","Jeffrey","Jeffrey","Madelyn","Cora","Anthony","Aria","Everly","Gabriella","Penelope","Scarlett","Jonathan","Patrick","Amy","Isabella","Brandon","Chloe","Michael","Richard","John","Karen","Justin","Kennedy","Amelia","Cora","Savannah","Joshua","Karen","Jessica","Lily","Nora","Alice","Sarah","Naomi","Emily","Karen","Violet","George","Elizabeth","David","Helen","Grace","Audrey","Caroline","Paul","Olivia","Stephen","Andrew","David","Kennedy","Madelyn","Brooklyn","John","Harper","Gary","Grace","Grace","Naomi","Donna","Ariana","Skylar","Isla","Deborah","Serenity","Claire","Ariana","Kennedy","Everly","Audrey","Sarah","Lily","Stella","Laura","Jennifer","George","Maya","Joseph","Nova","George","Chloe","Samantha","Gregory","Isabella","Nancy","Elizabeth","Riley","Avery","Claire","Elena","Eliana","Elizabeth","Michael","Gary","Allison","Richard","Kimberly","John","Donna","Evelyn","Jack","Michael","Paul","Ruth","Ariana","Violet","Scarlett","Victoria","Isabella","Larry","Autumn","Bella","Kevin","Joshua","Hazel","Violet","Allison","Paisley","Joshua","Timothy"]

# Randomization functions
def random_from_list(items):
    return random.choice(items) if items else ''

def random_int(min_val, max_val):
    return random.randint(min_val, max_val)

def random_date(start_date, end_date):
    delta = end_date - start_date
    random_days = random_int(0, int(delta.total_seconds() / 86400))
    return start_date + timedelta(days=random_days)

# Streamlit UI
st.title("PeopleTray Demo User Generator")
st.write("Enter parameters below to generate the import spreadsheet. Click 'Generate' to create and download the .xlsx file.")

# Center layout using columns (empty sides for centering)
left, center, right = st.columns([1, 2, 1])
with center:
    num_users = st.number_input("Number of users", min_value=1, max_value=500, value=10)
    roster_names_input = st.text_input("Roster names (comma-separated, or default 'Demo Roster')")
    site_ports_input = st.text_input("Site ports (comma-separated, or default 'Demo Site Port')")
    home_ports_input = st.text_input("Home/travel ports (comma-separated, or default 'Demo Home Port')")
    workgroups_input = st.text_input("Workgroups (comma-separated, required)")
    employment_types_input = st.text_input("Employment types (comma-separated, or default 'FullTime')")
    areas_input = st.text_input("Areas (comma-separated, or default 'Demo')")
    first_names_input = st.text_input("First names (comma-separated, or default 200-name list)")
    email_domain = st.text_input("Email domain (e.g., example.com)", value="example.com")
    work_roles_input = st.text_input("Work roles (comma-separated, or blank)")
    camp_port_mappings_input = st.text_input("Camp-port mappings ('Camp:Port', comma-separated, or blank)")
    starting_employee_id = st.text_input("Starting Employee ID (e.g., 00345, or blank)")

# Parse inputs with defaults
roster_names = [r.strip() for r in (roster_names_input or '').split(',') if r.strip()] or ['Demo Roster']
site_ports = [p.strip() for p in (site_ports_input or '').split(',') if p.strip()] or ['Demo Site Port']
home_ports = [p.strip() for p in (home_ports_input or '').split(',') if p.strip()] or ['Demo Home Port']
workgroups = [w.strip() for w in (workgroups_input or '').split(',') if w.strip()]
employment_types = [e.strip() for e in (employment_types_input or '').split(',') if e.strip()] or ['FullTime']
areas = [a.strip() for a in (areas_input or '').split(',') if a.strip()] or ['Demo']
work_roles = [w.strip() for w in (work_roles_input or '').split(',') if w.strip()] or ['']

# Parse camp-port mappings
camp_port_mappings = []
if camp_port_mappings_input.strip():
    for pair in camp_port_mappings_input.split(','):
        pair = pair.strip()
        if ':' in pair:
            camp, port = pair.split(':', 1)
            camp_port_mappings.append((camp.strip(), port.strip()))

# Use default first names or parse input
if first_names_input:
    first_names_list = [name.strip() for name in first_names_input.split(',') if name.strip()]

# Parse starting Employee ID (if provided, use for sequential generation)
starting_id = None
id_length = 0
if starting_employee_id.strip():
    try:
        starting_id = int(starting_employee_id.lstrip('0'))  # Remove leading zeros for calculation
        id_length = len(starting_employee_id)  # Preserve original length for formatting
    except ValueError:
        st.error("Invalid starting Employee ID. It must be numeric (e.g., 00345).")

# System options
salutations = ['Mr', 'Mrs', 'Ms', 'Miss', 'Dr']
genders = ['Male', 'Female', 'Not specified']

# Date range setup (current date: October 03, 2025)
today = datetime(2025, 10, 3)
five_years_ago = today - timedelta(days=5 * 365)
adult_start = datetime(1950, 1, 1)
adult_end = datetime(1990, 1, 1)

# Generate button (centered)
with center:
    if st.button("Generate Spreadsheet"):
        if not workgroups:
            st.error("Workgroups are required.")
        else:
            # Create workbook and sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PersonProfiles"

            # Populate rows 1-3 exactly as template
            for col, value in enumerate(row1, 1):
                ws.cell(row=1, column=col, value=value)
            for col, value in enumerate(row2, 1):
                ws.cell(row=2, column=col, value=value)
            for col, value in enumerate(row3, 1):
                ws.cell(row=3, column=col, value=value)

            # Bold headers (row 3, from column B onwards, A3 is 'DO NOT DELETE')
            header_font = Font(bold=True)
            for col in range(2, 85):  # Start from B3
                ws.cell(row=3, column=col).font = header_font

            # Generate data rows (starting row 4, column A blank, data from B)
            for user_index in range(num_users):
                row_num = 4 + user_index

                # Leave A blank
                ws.cell(row=row_num, column=1).value = ''

                # Core person details
                first_name = random_from_list(first_names_list)
                preferred_name = first_name
                middle_name = ''
                last_name = 'USER'
                email_suffix = random_int(100, 999)
                username = f"{first_name.lower()}.{last_name.lower()}{email_suffix}@{email_domain}"

                # Employee ID: Sequential if starting ID provided, else blank
                employee_id = ''
                if starting_id is not None:
                    current_id = starting_id + user_index
                    employee_id = f"{current_id:0{id_length}d}"  # Format with leading zeros

                salutation = random_from_list(salutations)
                gender = random_from_list(genders)
                dob = random_date(adult_start, adult_end)
                commenced = random_date(five_years_ago, today)
                date_ceased = ''

                # Work details
                workgroup = random_from_list(workgroups)
                work_role_01 = random_from_list(work_roles) if work_roles and work_roles[0] != '' else ''
                work_role_02 = ''  # Blank
                work_role_03 = ''  # Blank
                job_title = work_role_01  # Match Work Role 01
                area = random_from_list(areas)
                employment_type = random_from_list(employment_types)
                employer = ''  # Blank
                person_company = ''  # Blank

                # Contact details
                work_phone = '+61400000001'  # Fixed value
                work_mobile = '+61400000001'  # Fixed value
                additional_email = ''
                home_phone = ''
                home_mobile = ''
                home_email = ''

                # Address (residential and postal identical)
                address1 = '123 Demo Street'
                address2 = ''
                suburb = 'Demo Suburb'
                postcode = '6000'
                state = 'WA'
                country = 'Australia'
                postal_address1 = address1
                postal_address2 = address2
                postal_suburb = suburb
                postal_postcode = postcode
                postal_state = state
                postal_country = country

                # Emergency contacts (all blank)
                ec1_name = ''
                ec1_phone = ''
                ec1_email = ''
                ec1_relationship = ''
                ec1_guardian = ''
                ec1_can_speak = ''
                ec2_name = ''
                ec2_phone = ''
                ec2_email = ''
                ec2_relationship = ''
                ec2_guardian = ''
                ec2_can_speak = ''

                # Training (blanks)
                usi = ''
                default_rto = ''  # Blank

                # Roster details
                current_roster_name = random_from_list(roster_names)
                current_roster_from = random_date(five_years_ago, today)
                roster_seed_date = current_roster_from
                current_roster_to = ''  # Blank
                current_roster_notes = 'Generated demo roster notes'

                # Travel settings
                home_port = random_from_list(home_ports)
                default_site_port = random_from_list(site_ports)
                default_camp = ''  # Will be set from mappings if available
                default_block = ''  # Blank
                default_room_type = ''  # Blank
                default_room = ''  # Blank
                default_room_group = ''  # Blank
                default_cost_code = ''  # Blank

                # Set Default Camp and override Default Site Port if mappings provided
                if camp_port_mappings:
                    selected_camp, selected_port = random.choice(camp_port_mappings)
                    default_camp = selected_camp
                    default_site_port = selected_port

                # Travel options (blanks where specified)
                inbound_preference = ''  # Blank
                outbound_preference = ''  # Blank
                seat_preference = ''  # Blank
                weight_kg = random_int(60, 100)
                extra_baggage_kg = ''  # Blank

                # Dietary/Accessibility
                frequent_flyer1 = ''  # Blank
                frequent_flyer2 = ''  # Blank
                vegetarian = 'No'
                vegan = 'No'
                gluten_free = 'No'
                other_requirement = 'No'
                other_requirement_note = ''
                wheelchair_access_required = 'No'

                # Custom fields (blank)
                custom_field_01 = ''
                custom_field_02 = ''
                custom_field_03 = ''
                custom_field_04 = ''

                # Assemble row data (83 columns for B onwards)
                row_data = [
                    username, employee_id, salutation, last_name, first_name, middle_name, preferred_name, workgroup,
                    work_role_01, work_role_02, work_role_03, job_title, area, employment_type, employer, person_company,
                    gender, dob, commenced, date_ceased, work_phone, work_mobile, username, additional_email,
                    home_phone, home_mobile, home_email, address1, address2, suburb, postcode, state, country,
                    postal_address1, postal_address2, postal_suburb, postal_postcode, postal_state, postal_country,
                    ec1_name, ec1_phone, ec1_email, ec1_relationship, ec1_guardian, ec1_can_speak, ec2_name,
                    ec2_phone, ec2_email, ec2_relationship, ec2_guardian, ec2_can_speak, usi, default_rto, current_roster_name, current_roster_from,
                    roster_seed_date, current_roster_to, current_roster_notes, home_port, default_site_port,
                    default_camp, default_block, default_room_type, default_room, default_room_group, default_cost_code,
                    inbound_preference, outbound_preference, seat_preference, weight_kg, extra_baggage_kg, frequent_flyer1,
                    frequent_flyer2, vegetarian, vegan, gluten_free, other_requirement, other_requirement_note,
                    wheelchair_access_required, custom_field_01, custom_field_02, custom_field_03, custom_field_04
                ]

                # Write to worksheet starting from column 2 (B)
                for col, value in enumerate(row_data, 2):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    if isinstance(value, datetime):
                        cell.number_format = 'DD-MMM-YY'

            # Save to BytesIO for download
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            st.success(f"Generated spreadsheet with {num_users} users.")
            st.download_button(
                label="Download Generated_PeopleTray_Import.xlsx",
                data=output,
                file_name="Generated_PeopleTray_Import.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# Footer note
st.markdown("---")
st.write("*Delete rows 1–2 before importing to PeopleTray.*")
